import streamlit as st
import requests
import time
import datetime
import gzip
import json
import pandas as pd
import concurrent.futures
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- PAGE CONFIG ---
st.set_page_config(
    page_title="Shoeb Institutional Terminal",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Theme State Management
if "theme" not in st.session_state:
    st.session_state.theme = "Dark"

def toggle_theme():
    if st.session_state.theme == "Dark":
        st.session_state.theme = "Light"
    else:
        st.session_state.theme = "Dark"

# Colors for UI
if st.session_state.theme == "Dark":
    bg_color = "#07090e"
    sidebar_bg = "#0b0f17"
    panel_bg = "#0f172a"
    text_color = "#f8fafc"
    border_col = "#1e293b"
    table_bg = "#0b0f19"
    table_header = "#131c31"
    row_hover = "#162032"
else:
    bg_color = "#f1f5f9"
    sidebar_bg = "#ffffff"
    panel_bg = "#ffffff"
    text_color = "#0f172a"
    border_col = "#cbd5e1"
    table_bg = "#ffffff"
    table_header = "#e2e8f0"
    row_hover = "#f8fafc"

# --- PRO WEBSITE UI STYLING ---
st.markdown(f"""
<style>
    /* =========================================================
       SHOEB INSTITUTIONAL TERMINAL — PROFESSIONAL UI LAYER
       UI/CSS ONLY. Scanner/data/zone logic is untouched.
       ========================================================= */

    :root {{
        --bg: {bg_color};
        --sidebar: {sidebar_bg};
        --panel: {panel_bg};
        --text: {text_color};
        --border: {border_col};
        --muted: #94a3b8;
        --accent: #38bdf8;
        --accent-2: #6366f1;
        --success: #34d399;
        --danger: #f87171;
    }}

    /* ---------- GLOBAL APP ---------- */
    .stApp {{
        background:
            radial-gradient(circle at 78% 8%, rgba(56,189,248,.075), transparent 26%),
            radial-gradient(circle at 45% 90%, rgba(99,102,241,.055), transparent 30%),
            {bg_color};
        color: {text_color};
    }}

    [data-testid="stAppViewContainer"] > .main {{
        background: transparent;
    }}

    .main .block-container {{
        max-width: 1500px;
        padding-top: 2.4rem;
        padding-bottom: 3rem;
        padding-left: 2.4rem;
        padding-right: 2.4rem;
    }}

    #MainMenu,
    footer {{
        visibility: hidden;
    }}

    [data-testid="stHeader"] {{
        background: transparent;
        height: 0;
    }}

    [data-testid="stToolbar"] {{
        display: none;
    }}

    /* ---------- SIDEBAR ---------- */
    section[data-testid="stSidebar"] {{
        background:
            linear-gradient(180deg, rgba(15,23,42,.98) 0%, rgba(7,12,22,.99) 100%);
        border-right: 1px solid rgba(148,163,184,.12);
        box-shadow: 12px 0 35px rgba(0,0,0,.18);
    }}

    section[data-testid="stSidebar"] > div {{
        padding: 1.4rem 1rem 1.2rem;
    }}

    section[data-testid="stSidebar"] .block-container {{
        padding-top: 1rem;
    }}

    section[data-testid="stSidebar"] h3 {{
        color: #f8fafc !important;
        font-size: 1rem !important;
        font-weight: 800 !important;
        letter-spacing: .02em;
        margin-bottom: .5rem !important;
    }}

    section[data-testid="stSidebar"] hr {{
        border: 0;
        border-top: 1px solid rgba(148,163,184,.10);
        margin: 1.25rem 0;
    }}

    section[data-testid="stSidebar"] label {{
        color: #cbd5e1 !important;
        font-size: .82rem !important;
        font-weight: 600 !important;
    }}

    section[data-testid="stSidebar"] [data-baseweb="radio"] {{
        padding: .16rem 0;
    }}

    section[data-testid="stSidebar"] [data-baseweb="radio"] > div:first-child {{
        background: rgba(148,163,184,.10);
        border-color: rgba(148,163,184,.20);
    }}

    section[data-testid="stSidebar"] [data-baseweb="select"] > div {{
        background: rgba(15,23,42,.86);
        border: 1px solid rgba(148,163,184,.18);
        border-radius: 10px;
        min-height: 42px;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.025);
    }}

    section[data-testid="stSidebar"] [data-baseweb="select"] > div:hover {{
        border-color: rgba(56,189,248,.55);
    }}

    section[data-testid="stSidebar"] [data-testid="stNumberInput"] input {{
        background: rgba(15,23,42,.86) !important;
        color: #f8fafc !important;
        border: 1px solid rgba(148,163,184,.18) !important;
        border-radius: 10px !important;
    }}

    section[data-testid="stSidebar"] [data-testid="stCheckbox"] {{
        margin-top: .15rem;
    }}

    /* ---------- PREMIUM BUTTONS ---------- */
    section[data-testid="stSidebar"] .stButton > button {{
        width: 100%;
        min-height: 44px;
        border: 1px solid rgba(148,163,184,.18);
        border-radius: 11px;
        background: linear-gradient(135deg, rgba(30,41,59,.96), rgba(15,23,42,.96));
        color: #e2e8f0;
        font-weight: 700;
        letter-spacing: .01em;
        box-shadow: 0 8px 22px rgba(0,0,0,.18);
        transition: all .18s ease;
    }}

    section[data-testid="stSidebar"] .stButton > button:hover {{
        border-color: rgba(56,189,248,.55);
        color: #ffffff;
        transform: translateY(-1px);
        box-shadow: 0 10px 28px rgba(56,189,248,.12);
    }}

    section[data-testid="stSidebar"] button[kind="primary"] {{
        background: linear-gradient(135deg, #38bdf8 0%, #2563eb 100%) !important;
        border: 0 !important;
        color: white !important;
        min-height: 48px;
        font-size: .9rem;
        box-shadow: 0 10px 28px rgba(37,99,235,.28);
    }}

    section[data-testid="stSidebar"] button[kind="primary"]:hover {{
        background: linear-gradient(135deg, #60a5fa 0%, #4f46e5 100%) !important;
        transform: translateY(-1px);
        box-shadow: 0 14px 34px rgba(37,99,235,.35);
    }}

    /* ---------- MAIN HEADER ---------- */
    .terminal-header {{
        position: relative;
        overflow: hidden;
        padding: 28px 32px;
        border-radius: 20px;
        background:
            linear-gradient(135deg, rgba(15,23,42,.97), rgba(15,23,42,.78)),
            radial-gradient(circle at 90% 0%, rgba(56,189,248,.16), transparent 34%);
        border: 1px solid rgba(56,189,248,.16);
        margin-bottom: 26px;
        box-shadow:
            0 18px 50px rgba(0,0,0,.24),
            inset 0 1px 0 rgba(255,255,255,.035);
    }}

    .terminal-header::before {{
        content: "";
        position: absolute;
        left: 0;
        top: 0;
        width: 5px;
        height: 100%;
        background: linear-gradient(180deg, #38bdf8, #6366f1);
        border-radius: 20px 0 0 20px;
    }}

    .terminal-header::after {{
        content: "";
        position: absolute;
        width: 260px;
        height: 260px;
        right: -120px;
        top: -160px;
        border-radius: 50%;
        background: rgba(56,189,248,.08);
        filter: blur(2px);
    }}

    .terminal-header h1 {{
        position: relative;
        z-index: 1;
        margin: 0 !important;
        font-size: 26px !important;
        font-weight: 850 !important;
        letter-spacing: -.02em;
        background: linear-gradient(90deg, #f8fafc, #7dd3fc);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }}

    .terminal-header p {{
        position: relative;
        z-index: 1;
        color: #94a3b8 !important;
        font-size: 13px !important;
        margin-top: 8px !important;
        letter-spacing: .015em;
    }}

    /* ---------- METRIC CARDS ---------- */
    .premium-card {{
        position: relative;
        overflow: hidden;
        background:
            linear-gradient(145deg, rgba(15,23,42,.98), rgba(15,23,42,.78));
        border: 1px solid rgba(148,163,184,.13);
        padding: 20px 18px;
        min-height: 102px;
        border-radius: 16px;
        text-align: center;
        box-shadow:
            0 12px 32px rgba(0,0,0,.18),
            inset 0 1px 0 rgba(255,255,255,.025);
        margin-bottom: 15px;
        transition: transform .18s ease, border-color .18s ease;
    }}

    .premium-card:hover {{
        transform: translateY(-2px);
        border-color: rgba(56,189,248,.24);
    }}

    .premium-card::after {{
        content: "";
        position: absolute;
        left: 15%;
        right: 15%;
        bottom: -35px;
        height: 65px;
        background: rgba(56,189,248,.08);
        filter: blur(28px);
        border-radius: 50%;
    }}

    /* ---------- ALERTS / DOWNLOAD / PROGRESS ---------- */
    [data-testid="stAlert"] {{
        border-radius: 12px !important;
        border: 1px solid rgba(56,189,248,.16) !important;
        background: rgba(15,23,42,.72) !important;
    }}

    [data-testid="stDownloadButton"] > button {{
        border-radius: 11px !important;
        min-height: 44px !important;
        font-weight: 700 !important;
        border: 1px solid rgba(56,189,248,.22) !important;
        background: rgba(15,23,42,.86) !important;
        color: #e2e8f0 !important;
        transition: all .18s ease;
    }}

    [data-testid="stDownloadButton"] > button:hover {{
        border-color: rgba(56,189,248,.55) !important;
        color: #7dd3fc !important;
        transform: translateY(-1px);
    }}

    [data-testid="stProgress"] > div {{
        background: rgba(148,163,184,.10) !important;
        border-radius: 99px;
    }}

    [data-testid="stProgress"] > div > div {{
        background: linear-gradient(90deg, #38bdf8, #6366f1) !important;
        border-radius: 99px;
    }}

    /* Existing table: only visual polish; scrolling/sorting stay intact. */
    .shoeb-table-wrapper {{
        box-shadow:
            0 18px 45px rgba(0,0,0,.24),
            inset 0 1px 0 rgba(255,255,255,.025) !important;
        border-color: rgba(148,163,184,.14) !important;
    }}

    /* ---------- RESPONSIVE ---------- */
    @media (max-width: 900px) {{
        .main .block-container {{
            padding-left: 1rem;
            padding-right: 1rem;
        }}

        .terminal-header {{
            padding: 22px;
        }}

        .terminal-header h1 {{
            font-size: 21px !important;
        }}
    }}
</style>
""", unsafe_allow_html=True)

# --- PERMANENT HARDCODED TOKEN ---
GLOBAL_TOKEN = "eyJ0eXAiOiJKV1QiLCJrZXlfaWQiOiJza192MS4wIiwiYWxnIjoiSFMyNTYifQ.eyJzdWIiOiJIVTU0OTgiLCJqdGkiOiI2YTc5OGU2ZTBmZDM2ODI2MDg5NTJiNmQiLCJpc011bHRpQ2xpZW50IjpmYWxzZSwiaXNQbHVzUGxhbiI6dHJ1ZSwiaXNFeHRlbmRlZCI6dHJ1ZSwiaWF0IjoxNzg2MzUxMjE0LCJpc3MiOiJ1ZGFwaS1nYXRld2F5LXNlcnZpY2UiLCJleHAiOjE4MTc5MzUyMDB9.yXDWPFJwNAyTlQGzf_olWHdxxwG6q4blw1j0037WjQs"

@st.cache_data
def get_nifty_500_instruments():
    """
    Load the ACTUAL NIFTY 500 constituents from the official NSE/NSE Indices CSV,
    then map those 500 constituents to Upstox NSE_EQ instrument keys by ISIN.

    IMPORTANT:
    - Replaces the old incorrect equity_keys[:150] logic.
    - No Demand/Supply, zone, HTF, RVOL, table, CSV or Excel logic is changed.
    - The web app stops instead of silently scanning an incomplete/wrong universe.
    """

    # ---------------------------------------------------------
    # 1) Load official NIFTY 500 constituent list
    # ---------------------------------------------------------
    nifty500_urls = [
        "https://nsearchives.nseindia.com/content/indices/ind_nifty500list.csv",
        "https://www.niftyindices.com/IndexConstituent/ind_nifty500list.csv",
    ]

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/151.0.0.0 Safari/537.36"
        ),
        "Accept": "text/csv,application/csv,text/plain,*/*",
        "Referer": "https://www.nseindia.com/",
    }

    nifty500_rows = None
    last_error = None

    for csv_url in nifty500_urls:
        try:
            response = requests.get(csv_url, headers=headers, timeout=20)
            response.raise_for_status()

            text = response.content.decode("utf-8-sig", errors="replace")
            reader = pd.read_csv(BytesIO(response.content))

            if reader.empty:
                raise ValueError("NIFTY 500 CSV returned no rows.")

            # Normalize column names.
            reader.columns = [
                str(col).strip().lower().replace(" ", "_")
                for col in reader.columns
            ]

            symbol_key = next(
                (col for col in reader.columns if col in ("symbol", "symbols")),
                None
            )

            isin_key = next(
                (
                    col
                    for col in reader.columns
                    if col in ("isin_code", "isin", "isin_code_")
                ),
                None
            )

            if not symbol_key:
                raise ValueError(
                    f"Could not find Symbol column. Columns: {list(reader.columns)}"
                )

            if not isin_key:
                raise ValueError(
                    f"Could not find ISIN Code column. Columns: {list(reader.columns)}"
                )

            nifty500_rows = []

            for _, row in reader.iterrows():
                symbol = str(row.get(symbol_key, "") or "").strip().upper()
                isin = str(row.get(isin_key, "") or "").strip().upper()

                if symbol:
                    nifty500_rows.append({
                        "symbol": symbol,
                        "isin": isin
                    })

            break

        except Exception as e:
            last_error = e

    if not nifty500_rows:
        raise RuntimeError(
            "Could not load the official NIFTY 500 constituent list. "
            f"Scanner stopped to prevent an incorrect universe. Last error: {last_error}"
        )

    # ---------------------------------------------------------
    # 2) Remove duplicate constituents
    # ---------------------------------------------------------
    unique_constituents = []
    seen_isins = set()
    seen_symbols = set()

    for item in nifty500_rows:
        isin = item["isin"]
        symbol = item["symbol"]

        if isin and isin not in seen_isins:
            unique_constituents.append(item)
            seen_isins.add(isin)
        elif not isin and symbol and symbol not in seen_symbols:
            unique_constituents.append(item)
            seen_symbols.add(symbol)

    if len(unique_constituents) != 500:
        raise RuntimeError(
            f"Expected exactly 500 NIFTY 500 constituents, "
            f"but received {len(unique_constituents)}. "
            "Scanner stopped to prevent an incomplete/wrong universe."
        )

    # ---------------------------------------------------------
    # 3) Download Upstox master instrument list
    # ---------------------------------------------------------
    upstox_url = "https://assets.upstox.com/market-quote/instruments/exchange/NSE.json.gz"

    response = requests.get(upstox_url, timeout=30)
    response.raise_for_status()

    decompressed_content = gzip.decompress(response.content)
    instruments = json.loads(decompressed_content.decode("utf-8"))

    # ---------------------------------------------------------
    # 4) Build reliable ISIN + symbol mappings
    # ---------------------------------------------------------
    isin_to_key = {}
    symbol_to_key = {}

    for inst in instruments:
        if (
            inst.get("segment") == "NSE_EQ"
            and inst.get("instrument_type") in {"EQ", "BE"}
            and inst.get("instrument_key")
        ):
            instrument_key = inst.get("instrument_key")

            isin = str(inst.get("isin") or "").strip().upper()

            trading_symbol = str(
                inst.get("trading_symbol")
                or inst.get("symbol")
                or ""
            ).strip().upper()

            if isin:
                isin_to_key[isin] = instrument_key

            if trading_symbol:
                symbol_to_key[trading_symbol] = instrument_key

    # ---------------------------------------------------------
    # 5) Map all official NIFTY 500 constituents to Upstox keys
    # ---------------------------------------------------------
    mapped_instruments = []
    missing = []

    for constituent in unique_constituents:
        symbol = constituent["symbol"]
        isin = constituent["isin"]

        instrument_key = None

        # Primary mapping: ISIN
        if isin:
            instrument_key = isin_to_key.get(isin)

        # Fallback: exact trading symbol
        if not instrument_key:
            instrument_key = symbol_to_key.get(symbol)

        if instrument_key:
            mapped_instruments.append((symbol, instrument_key))
        else:
            missing.append(f"{symbol} ({isin})")

    # Remove duplicate instrument keys while preserving order.
    seen_keys = set()
    final_instruments = []

    for symbol, instrument_key in mapped_instruments:
        if instrument_key not in seen_keys:
            final_instruments.append((symbol, instrument_key))
            seen_keys.add(instrument_key)

    # ---------------------------------------------------------
    # 6) Strict verification — never silently scan less than 500
    # ---------------------------------------------------------
    if missing:
        print("\n❌ Missing NIFTY 500 -> Upstox mappings:")
        for item in missing:
            print(f"   - {item}")

    print(f"✅ Official NIFTY 500 stocks: {len(unique_constituents)}")
    print(f"✅ Upstox instruments mapped: {len(final_instruments)}")
    print(f"⚠️ Missing Upstox mappings: {len(missing)}")

    if len(final_instruments) != 500:
        raise RuntimeError(
            f"NIFTY 500 mapping incomplete: "
            f"{len(final_instruments)}/500 stocks mapped. "
            "Scanner stopped instead of scanning an incorrect universe."
        )

    # ---------------------------------------------------------
    # 7) Explicit RELIANCE verification
    # ---------------------------------------------------------
    reliance_key = None

    for symbol, instrument_key in final_instruments:
        if symbol == "RELIANCE":
            reliance_key = instrument_key
            break

    if not reliance_key:
        raise RuntimeError(
            "RELIANCE could not be verified in the NIFTY 500 -> "
            "Upstox mapping. Scanner stopped."
        )

    print(f"✅ RELIANCE verified -> {reliance_key}")
    print("🚀 NIFTY 500 universe verification complete: 500/500")

    return final_instruments


def get_upstox_historical_candles(instrument_key, interval):
    to_date = datetime.date.today().isoformat()
    
    api_interval = interval
    is_resampled = False
    resample_rule = ""

    if "minute" in interval:
        mins = int(interval.replace("minute", ""))
        if mins > 1 and mins not in [5, 30]:
            api_interval = "1minute"
            is_resampled = True
            resample_rule = f"{mins}min"
            from_date = (datetime.date.today() - datetime.timedelta(days=15)).isoformat()
        else:
            from_date = (datetime.date.today() - datetime.timedelta(days=10)).isoformat()
    elif interval in ["30minute", "60minute", "day"]:
        from_date = (datetime.date.today() - datetime.timedelta(days=180)).isoformat()
    else:
        from_date = (datetime.date.today() - datetime.timedelta(days=730)).isoformat()
        
    url = f"https://api.upstox.com/v2/historical-candle/{instrument_key}/{api_interval}/{to_date}/{from_date}"
    headers = {'Accept': 'application/json', 'Authorization': f'Bearer {GLOBAL_TOKEN}'}
    
    try:
        time.sleep(0.12)
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if 'data' in data and data['data'] and 'candles' in data['data']:
                raw_candles = data['data']['candles']
                
                if is_resampled:
                    df = pd.DataFrame(raw_candles, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'oi'])
                    df['timestamp'] = pd.to_datetime(df['timestamp'])
                    df.set_index('timestamp', inplace=True)
                    df = df.sort_index()
                    
                    resampled = df.resample(resample_rule).agg({
                        'open': 'first',
                        'high': 'max',
                        'low': 'min',
                        'close': 'last',
                        'volume': 'sum'
                    }).dropna()
                    
                    return resampled[['open', 'high', 'low', 'close', 'volume']].values.tolist()
                else:
                    return [[float(c[1]), float(c[2]), float(c[3]), float(c[4]), float(c[5])] for c in reversed(raw_candles)]
    except:
        pass
    return []


def classify_candle(open_price, high_price, low_price, close_price):
    body_size = abs(close_price - open_price)
    total_range = high_price - low_price
    if total_range == 0: 
        return "Doji", 0
    
    body_percentage = (body_size / total_range) * 100
    is_green = close_price > open_price
    
    if body_percentage <= 40:
        return "Base Candle", body_percentage
    else:
        return ("Green Strong" if is_green else "Red Strong"), body_percentage


def extract_zones(candles_data, scan_type, marking_style, cmp_price, is_higher_tf=False):
    zones = []
    if not candles_data or len(candles_data) < 7:
        return zones

    i = 0
    n = len(candles_data)
    
    while i < n - 2:
        leg_in = candles_data[i]
        t_in, _ = classify_candle(leg_in[0], leg_in[1], leg_in[2], leg_in[3])
        
        base_candles = []
        j = i + 1
        while j < n - 1:
            b = candles_data[j]
            b_type, _ = classify_candle(b[0], b[1], b[2], b[3])
            if b_type == "Base Candle":
                base_candles.append(b)
                j += 1
            else:
                break
        
        if len(base_candles) > 0 and len(base_candles) <= 5:
            leg_out_idx = i + 1 + len(base_candles)
            if leg_out_idx < n:
                leg_out = candles_data[leg_out_idx]
                t_out, _ = classify_candle(leg_out[0], leg_out[1], leg_out[2], leg_out[3])
                
                target_base_count = len(base_candles)
                
                # --- DEMAND ZONE ---
                if scan_type in ["Demand & Supply", "Demand Zone Only"] and ("Green" in t_out) and (leg_out[3] > leg_in[3]):
                    if marking_style == "Body to Wick":
                        proximal = max(max(b[0], b[3]) for b in base_candles)
                    else: 
                        proximal = max(b[1] for b in base_candles)
                        
                    distal = min(b[2] for b in base_candles)
                    
                    if proximal < cmp_price or is_higher_tf:
                        is_tested = False
                        zone_index = i + 1
                        
                        for future_idx in range(leg_out_idx + 1, n):
                            f_candle = candles_data[future_idx]
                            if f_candle[2] <= proximal:
                                is_tested = True
                                break
                        
                        if not is_tested:
                            pattern = "DBR" if "Red" in t_in else "RBR"
                            zones.append({
                                "Index": zone_index,
                                "Zone": "Demand Zone",
                                "Pattern": pattern,
                                "Base Group Size": target_base_count,
                                "Proximal": round(proximal, 2),
                                "Distal": round(distal, 2)
                            })
                
                # --- SUPPLY ZONE ---
                elif scan_type in ["Demand & Supply", "Supply Zone Only"] and ("Red" in t_out) and (leg_out[3] < leg_in[3]):
                    if marking_style == "Body to Wick":
                        proximal = min(min(b[0], b[3]) for b in base_candles)
                    else: 
                        proximal = min(b[2] for b in base_candles)
                        
                    distal = max(b[1] for b in base_candles)
                    
                    if proximal > cmp_price or is_higher_tf:
                        is_tested = False
                        zone_index = i + 1
                        
                        for future_idx in range(leg_out_idx + 1, n):
                            f_candle = candles_data[future_idx]
                            if f_candle[1] >= proximal:
                                is_tested = True
                                break
                        
                        if not is_tested:
                            pattern = "RBD" if "Green" in t_in else "DBD"
                            zones.append({
                                "Index": zone_index,
                                "Zone": "Supply Zone",
                                "Pattern": pattern,
                                "Base Group Size": target_base_count,
                                "Proximal": round(proximal, 2),
                                "Distal": round(distal, 2)
                            })
                
                i = leg_out_idx
                continue
                
        i += 1
        
    return zones


def scan_stock(symbol, key, interval, scan_type, marking_style):
    results = []
    raw_data = get_upstox_historical_candles(key, interval)
    if not raw_data or len(raw_data) < 7:
        return results

    ltf_candles = [[c[0], c[1], c[2], c[3]] for c in raw_data]
    volumes = [c[4] for c in raw_data]

    cmp_price = ltf_candles[-1][3]
    prev_close = ltf_candles[-2][3]
    change_pct = ((cmp_price - prev_close) / prev_close) * 100

    vol_series = pd.Series(volumes)
    if len(vol_series) >= 21:
        avg_vol_20 = vol_series.iloc[-21:-1].mean()
        current_vol = volumes[-1]
        rvol = current_vol / avg_vol_20 if avg_vol_20 > 0 else 1.0
    else:
        rvol = 1.0

    ltf_zones = extract_zones(ltf_candles, scan_type, marking_style, cmp_price, is_higher_tf=False)

    htf_map = {
        "1minute": "5minute", "5minute": "30minute", "15minute": "60minute",
        "30minute": "day", "60minute": "day", "day": "week", "week": "month", "month": "month"
    }
    htf_interval = htf_map.get(interval, "day")
    htf_raw = get_upstox_historical_candles(key, htf_interval)
    htf_candles = [[c[0], c[1], c[2], c[3]] for c in htf_raw] if htf_raw else []
    htf_zones = extract_zones(htf_candles, scan_type, marking_style, cmp_price, is_higher_tf=True) if htf_candles else []

    stock_zones = []
    for z in ltf_zones:
        htf_aligned = False
        for hz in htf_zones:
            if z["Zone"] == hz["Zone"]:
                if not (z["Proximal"] < hz["Distal"] or z["Distal"] > hz["Proximal"]):
                    htf_aligned = True
                    break
        
        htf_status = "🟢 HTF Aligned (A+)" if htf_aligned else "⚪ Local Zone Only"
        
        stock_zones.append({
            "Index": z["Index"],
            "Symbol": symbol,
            "Zone": z["Zone"],
            "Pattern": z["Pattern"],
            "Base Size": z["Base Group Size"],
            "Timeframe": interval,
            "HTF Status": htf_status,
            "Proximal": z["Proximal"],
            "Distal": z["Distal"],
            "CMP": cmp_price,
            "Distance (%)": round(
                (abs(z["Proximal"] - cmp_price) / cmp_price) * 100, 2
            ) if cmp_price > 0 else 0.0,
            "Change (%)": round(change_pct, 2),
            "RVOL": round(rvol, 2),
            "DistanceToCMP": abs(z["Proximal"] - cmp_price)
        })

    stock_zones = sorted(stock_zones, key=lambda x: x["DistanceToCMP"])
    
    unique_zones = []
    seen_ranges = set()
    for z in stock_zones:
        range_signature = (z["Proximal"], z["Distal"])
        if range_signature not in seen_ranges:
            seen_ranges.add(range_signature)
            unique_zones.append(z)

    for z in unique_zones[:5]:
        z_copy = z.copy()
        del z_copy["Index"]
        del z_copy["DistanceToCMP"]
        results.append(z_copy)
        
    return results


# --- STREAMLIT UI ---
st.markdown("""
    <div class="terminal-header">
        <h1 style="margin:0; font-size: 24px; font-weight: 800; color: #38bdf8; letter-spacing: 0.5px;">⚡ Shoeb Institutional Terminal</h1>
        <p style="margin:5px 0 0 0; font-size: 13px; opacity: 0.75;">Smart Money Concept (SMC) & Multi-Timeframe Confluence Engine</p>
    </div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("<h3 style='margin-bottom:0;'>⚙️ Terminal Control</h3>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    
    if st.button("🌓 Toggle Day/Night Mode", use_container_width=True):
        toggle_theme()
        st.rerun()
        
    st.markdown(f"<p style='text-align:center; font-size:12px; color:gray; margin-top:10px;'>Theme: <b>{st.session_state.theme} Mode</b></p>", unsafe_allow_html=True)
    st.markdown("---")
    
    scan_type = st.radio("What to Scan?", ["Demand & Supply", "Demand Zone Only", "Supply Zone Only"])
    st.markdown("<br>", unsafe_allow_html=True)
    
    marking_style = st.radio("Zone Marking Style", ["Body to Wick", "Wick to Wick"])
    st.markdown("<br>", unsafe_allow_html=True)
    
    tf_options = ["1minute", "5minute", "15minute", "30minute", "60minute", "day", "week", "month", "Custom Minute"]
    selected_tf_choice = st.selectbox("Select Timeframe", tf_options, index=5)
    
    if selected_tf_choice == "Custom Minute":
        custom_mins = st.number_input("Enter Minutes (e.g., 3, 10, 45, 240)", min_value=1, max_value=1440, value=15)
        selected_tf = f"{custom_mins}minute"
    else:
        selected_tf = selected_tf_choice
        
    st.markdown("<br>", unsafe_allow_html=True)
    only_htf = st.checkbox("🎯 Show Only HTF Aligned (A+) Zones", value=False)
    st.markdown("<br>", unsafe_allow_html=True)

    max_distance_pct = st.slider(
        "📍 Max Proximal Distance from CMP (%)",
        min_value=0.5,
        max_value=20.0,
        value=5.0,
        step=0.5,
        help="Show only zones whose Proximal is within this percentage of CMP."
    )
    st.markdown("<br>", unsafe_allow_html=True)
    
    scan_button = st.button("🚀 Run Live Scan", use_container_width=True, type="primary")

if scan_button:
    st.info(f"Scanning Nifty 500 universe for {selected_tf} timeframe with professional engine...")
    instruments = get_nifty_500_instruments()
    all_zones = []
    
    progress_bar = st.progress(0)
    total = len(instruments)
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(scan_stock, sym, key, selected_tf, scan_type, marking_style): sym for sym, key in instruments}
        
        completed = 0
        for future in concurrent.futures.as_completed(futures):
            completed += 1
            progress_bar.progress(completed / total)
            try:
                res = future.result()
                if res:
                    filtered_res = [
                        r for r in res
                        if r["Distance (%)"] <= max_distance_pct
                    ]

                    if only_htf:
                        filtered_res = [
                            r for r in filtered_res
                            if "HTF Aligned" in r["HTF Status"]
                        ]

                    all_zones.extend(filtered_res)
            except:
                pass
                
    st.success(f"Scan Complete! Successfully mapped {len(all_zones)} institutional zones.")
    
    if all_zones:
        df = pd.DataFrame(all_zones)
        
        # Metric Cards Layout
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f'<div class="premium-card"><h4 style="font-size:12px; color:gray; margin-bottom:4px; text-transform:uppercase; letter-spacing:0.5px;">Total Zones Found</h4><h2 style="color:#38bdf8; margin:0; font-weight:700;">{len(df)}</h2></div>', unsafe_allow_html=True)
        with col2:
            a_plus_count = len(df[df["HTF Status"].str.contains("A+")])
            st.markdown(f'<div class="premium-card"><h4 style="font-size:12px; color:gray; margin-bottom:4px; text-transform:uppercase; letter-spacing:0.5px;">A+ HTF Confluence</h4><h2 style="color:#34d399; margin:0; font-weight:700;">{a_plus_count}</h2></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="premium-card"><h4 style="font-size:12px; color:gray; margin-bottom:4px; text-transform:uppercase; letter-spacing:0.5px;">Active Timeframe</h4><h2 style="color:#f43f5e; margin:0; font-weight:700;">{selected_tf.upper()}</h2></div>', unsafe_allow_html=True)
            
        st.markdown("<br>", unsafe_allow_html=True)
        
        # --- RESPONSIVE SCREEN-FITTED PRO TERMINAL TABLE RENDERER ---
        table_html = f"""
        <html>
        <head>
        <style>
            body {{
                background-color: {table_bg};
                color: {text_color};
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
                margin: 0;
                padding: 0;
            }}
            .shoeb-table-wrapper {{
                width: 100%;
                max-height: 600px;
                background: {table_bg};
                border: 1px solid {border_col};
                border-radius: 12px;
                box-shadow: 0 10px 30px -5px rgba(0, 0, 0, 0.3);
                overflow-y: auto;
                overflow-x: auto;
            }}
            .shoeb-pro-table {{
                width: 100%;
                border-collapse: collapse;
                table-layout: auto;
            }}
            .shoeb-pro-table th {{
                background-color: {table_header};
                color: {text_color};
                padding: 12px 8px;
                font-size: 10.5px;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                border-bottom: 2px solid {border_col};
                overflow: hidden;
                text-overflow: ellipsis;
                white-space: nowrap;
                position: sticky;
                top: 0;
                z-index: 2;
            }}

            .change-sort-header {{
                cursor: pointer;
                user-select: none;
            }}

            .change-sort-header:hover {{
                color: #38bdf8;
            }}

            .sort-indicator {{
                font-size: 11px;
                margin-left: 3px;
            }}
            .shoeb-pro-table td {{
                padding: 11px 8px;
                color: {text_color};
                font-size: 12px;
                border-bottom: 1px solid {border_col};
                overflow: hidden;
                text-overflow: ellipsis;
                white-space: nowrap;
            }}

            .symbol-column {{
                width: 1%;
                white-space: nowrap !important;
                overflow: visible !important;
                text-overflow: clip !important;
            }}

            .base-column {{
                width: 1%;
                white-space: nowrap !important;
                text-align: center;
            }}
            .shoeb-pro-table tr:hover {{
                background-color: {row_hover};
            }}
            .badge-demand {{
                background-color: rgba(52, 211, 153, 0.12);
                color: #34d399;
                padding: 3px 6px;
                border-radius: 5px;
                font-weight: 600;
                font-size: 10px;
                border: 1px solid rgba(52, 211, 153, 0.25);
            }}
            .badge-supply {{
                background-color: rgba(248, 113, 113, 0.12);
                color: #f87171;
                padding: 3px 6px;
                border-radius: 5px;
                font-weight: 600;
                font-size: 10px;
                border: 1px solid rgba(248, 113, 113, 0.25);
            }}
            .badge-htf {{
                background-color: rgba(56, 189, 248, 0.12);
                color: #38bdf8;
                padding: 3px 5px;
                border-radius: 5px;
                font-weight: 600;
                font-size: 9.5px;
                border: 1px solid rgba(56, 189, 248, 0.25);
            }}
        </style>
        </head>
        <body>
        <div class="shoeb-table-wrapper">
            <table class="shoeb-pro-table">
                <thead>
                    <tr>
                        <th class="symbol-column">Symbol</th>
                        <th style="width: 11%;">Zone Type</th>
                        <th style="width: 7%;">Pattern</th>
                        <th class="base-column">Base</th>
                        <th style="width: 8%;">TF</th>
                        <th style="width: 13%;">HTF Status</th>
                        <th style="text-align:right; width: 9%;">Proximal</th>
                        <th style="text-align:right; width: 9%;">Distal</th>
                        <th style="text-align:right; width: 8%;">CMP</th>
                        <th class="distance-sort-header" onclick="sortDistancePercent()" style="text-align:right; width: 9%; cursor:pointer;" title="Click to sort Distance % ascending / descending">Distance % <span id="distanceSortIndicator" class="sort-indicator">↕</span></th>
                        <th class="change-sort-header" onclick="sortChangePercent()" style="text-align:right; width: 10%; cursor:pointer;" title="Click to sort Change % ascending / descending">Change % <span id="changeSortIndicator" class="sort-indicator">↕</span></th>
                        <th style="text-align:right; width: 9%;">RVOL</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for _, row in df.iterrows():
            zone_badge = f"<span class='badge-demand'>{row['Zone']}</span>" if "Demand" in row['Zone'] else f"<span class='badge-supply'>{row['Zone']}</span>"
            htf_badge = f"<span class='badge-htf'>{row['HTF Status']}</span>" if "A+" in row['HTF Status'] else f"<span style='color:gray; font-size:10px;'>{row['HTF Status']}</span>"
            
            chg_color = "#34d399" if row['Change (%)'] > 0 else ("#f87171" if row['Change (%)'] < 0 else "#f8fafc")
            
            table_html += f"""
                    <tr>
                        <td class="symbol-column" style="font-weight:700; color:#38bdf8;" title="{row['Symbol']}">{row['Symbol']}</td>
                        <td>{zone_badge}</td>
                        <td style="font-weight:600;">{row['Pattern']}</td>
                        <td class="base-column" style="font-weight:600;">{row['Base Size']}</td>
                        <td title="{row['Timeframe']}">{row['Timeframe']}</td>
                        <td>{htf_badge}</td>
                        <td style="text-align:right; font-weight:600; color:#34d399;">₹ {row['Proximal']:,.2f}</td>
                        <td style="text-align:right; font-weight:600; color:#f87171;">₹ {row['Distal']:,.2f}</td>
                        <td style="text-align:right; font-weight:600;">₹ {row['CMP']:,.2f}</td>
                        <td style="text-align:right; font-weight:700;">{row['Distance (%)']:.2f}%</td>
                        <td style="text-align:right; font-weight:600; color:{chg_color};">{row['Change (%)']:+.2f}%</td>
                        <td style="text-align:right; font-weight:600;">{row['RVOL']:.2f}x</td>
                    </tr>
            """
            
        table_html += """
                </tbody>
            </table>
        </div>

        <script>
            function getSortDirection(table, key) {
                return table.getAttribute('data-' + key + '-sort-direction') || 'none';
            }

            function setSortDirection(table, key, direction) {
                table.setAttribute('data-' + key + '-sort-direction', direction);
            }

            function sortTableByColumn(columnIndex, key, indicatorId) {
                const table = document.querySelector('.shoeb-pro-table');
                const tbody = table.tBodies[0];
                const rows = Array.from(tbody.rows);

                const currentDirection = getSortDirection(table, key);
                const newDirection = currentDirection === 'asc' ? 'desc' : 'asc';

                rows.sort(function(a, b) {
                    const aValue = parseFloat(a.cells[columnIndex].innerText.replace('%', '').replace('+', '').trim()) || 0;
                    const bValue = parseFloat(b.cells[columnIndex].innerText.replace('%', '').replace('+', '').trim()) || 0;

                    return newDirection === 'asc'
                        ? aValue - bValue
                        : bValue - aValue;
                });

                rows.forEach(function(row) {
                    tbody.appendChild(row);
                });

                setSortDirection(table, key, newDirection);

                const indicator = document.getElementById(indicatorId);
                if (indicator) {
                    indicator.textContent = newDirection === 'asc' ? '↑' : '↓';
                }
            }

            function sortDistancePercent() {
                sortTableByColumn(9, 'distance', 'distanceSortIndicator');
            }

            function sortChangePercent() {
                sortTableByColumn(10, 'change', 'changeSortIndicator');
            }
        </script>

        </body>
        </html>
        """

        st.components.v1.html(table_html, height=620, scrolling=False)
        
        st.markdown("<br>", unsafe_allow_html=True)
        csv_columns = [
            "Symbol", "Zone", "Pattern", "Base Size", "Timeframe",
            "HTF Status", "Proximal", "Distal", "CMP", "Distance (%)", "Change (%)", "RVOL"
        ]

        # ---------------------------------------------------------
        # RAW CSV EXPORT — existing CSV functionality preserved
        # ---------------------------------------------------------
        report_df = df[csv_columns].copy()
        csv = report_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')

        col_csv, col_excel = st.columns(2)

        with col_csv:
            st.download_button(
                "📥 Download CSV",
                data=csv,
                file_name=f"institutional_zones_{selected_tf}.csv",
                mime="text/csv",
                use_container_width=True
            )

        # ---------------------------------------------------------
        # PROFESSIONAL EXCEL INSTITUTIONAL REPORT
        # This changes only the export/report layer.
        # Scanner, zone detection, API and table logic are untouched.
        # ---------------------------------------------------------
        excel_buffer = BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            report_df.to_excel(
                writer,
                index=False,
                sheet_name="Institutional Report",
                startrow=3
            )

        excel_buffer.seek(0)
        workbook = load_workbook(excel_buffer)
        worksheet = workbook["Institutional Report"]

        # Professional title area
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(csv_columns))
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "⚡ SHOEB INSTITUTIONAL TERMINAL — INSTITUTIONAL ZONE REPORT"
        title_cell.font = Font(
            name="Calibri",
            size=16,
            bold=True,
            color="FFFFFF"
        )
        title_cell.fill = PatternFill("solid", fgColor="0F172A")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 30

        # Report metadata
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(csv_columns))
        meta_cell = worksheet.cell(row=2, column=1)
        meta_cell.value = f"Timeframe: {selected_tf.upper()}   |   Total Zones: {len(report_df)}"
        meta_cell.font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color="CBD5E1"
        )
        meta_cell.fill = PatternFill("solid", fgColor="111827")
        meta_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[2].height = 21

        # Blank separator row
        for cell in worksheet[3]:
            cell.fill = PatternFill("solid", fgColor="FFFFFF")

        # Header row is row 4
        header_fill = PatternFill("solid", fgColor="172554")
        header_font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color="FFFFFF"
        )
        thin_border = Border(
            bottom=Side(style="thin", color="CBD5E1")
        )

        for cell in worksheet[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = thin_border

        worksheet.row_dimensions[4].height = 24

        # Table range
        last_row = 4 + len(report_df)
        last_col_letter = chr(64 + len(csv_columns))

        if len(report_df) > 0:
            table_ref = f"A4:{last_col_letter}{last_row}"
            excel_table = Table(
                displayName="InstitutionalZoneReport",
                ref=table_ref
            )
            table_style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            excel_table.tableStyleInfo = table_style
            worksheet.add_table(excel_table)

        # Specific professional widths.
        # Do not iterate over worksheet.columns here because the title/header
        # area contains merged cells, which can produce MergedCell objects.
        # Fixed content-aware widths also keep the report stable in Excel.
        widths = {
            "A": 18,   # Symbol
            "B": 18,   # Zone
            "C": 12,   # Pattern
            "D": 10,   # Base Size
            "E": 13,   # Timeframe
            "F": 25,   # HTF Status
            "G": 14,   # Proximal
            "H": 14,   # Distal
            "I": 14,   # CMP
            "J": 14,   # Distance %
            "K": 14,   # Change %
            "L": 10    # RVOL
        }

        for col_letter, width in widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Freeze title/header area and enable Excel filtering via table.
        worksheet.freeze_panes = "A5"
        worksheet.sheet_view.showGridLines = False
        worksheet.auto_filter.ref = f"A4:{last_col_letter}{last_row}"

        # Number formatting and alignment
        for row in range(5, last_row + 1):
            worksheet.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            worksheet.cell(row=row, column=2).alignment = Alignment(horizontal="left")
            worksheet.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            worksheet.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            worksheet.cell(row=row, column=5).alignment = Alignment(horizontal="center")
            worksheet.cell(row=row, column=6).alignment = Alignment(horizontal="left")

            for col in [7, 8, 9]:
                worksheet.cell(row=row, column=col).number_format = '₹ #,##0.00'
                worksheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")

            # Distance (%) and Change (%) are stored as 19.98, not 0.1998.
            worksheet.cell(row=row, column=10).number_format = '0.00"%"'
            worksheet.cell(row=row, column=10).alignment = Alignment(horizontal="right")

            worksheet.cell(row=row, column=11).number_format = '+0.00"%" ; -0.00"%" ; 0.00"%"'
            worksheet.cell(row=row, column=11).alignment = Alignment(horizontal="right")

            worksheet.cell(row=row, column=12).number_format = '0.00"x"'
            worksheet.cell(row=row, column=12).alignment = Alignment(horizontal="right")

        # Positive/negative Change % highlighting
        green_fill = PatternFill("solid", fgColor="DCFCE7")
        green_font = Font(color="166534", bold=True)
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        red_font = Font(color="991B1B", bold=True)

        if len(report_df) > 0:
            change_range = f"K5:K{last_row}"
            worksheet.conditional_formatting.add(
                change_range,
                CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=green_fill,
                    font=green_font
                )
            )
            worksheet.conditional_formatting.add(
                change_range,
                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=red_fill,
                    font=red_font
                )
            )

        # Zone-specific font emphasis
        for row in range(5, last_row + 1):
            zone_cell = worksheet.cell(row=row, column=2)
            zone_text = str(zone_cell.value or "")

            if "Demand" in zone_text:
                zone_cell.font = Font(
                    name="Calibri",
                    size=10,
                    bold=True,
                    color="047857"
                )
            elif "Supply" in zone_text:
                zone_cell.font = Font(
                    name="Calibri",
                    size=10,
                    bold=True,
                    color="B91C1C"
                )

        # Workbook properties
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.orientation = "landscape"

        final_excel_buffer = BytesIO()
        workbook.save(final_excel_buffer)
        final_excel_buffer.seek(0)

        with col_excel:
            st.download_button(
                "📊 Download Professional Excel Report",
                data=final_excel_buffer.getvalue(),
                file_name=f"institutional_zones_{selected_tf}_professional.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.warning("No valid zones found matching your active filters.")